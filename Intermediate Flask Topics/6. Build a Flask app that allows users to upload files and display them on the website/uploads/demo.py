'''import pandas as pd
df=pd.read_csv('Room Allocation 10th 11th 12th 9th2023-10-06.csv')
header = pd.DataFrame({'Header 1': ['Value 1'], 'Header 2': ['Value 2']})
total=pd.pivot_table(df,index=['Class','Sub_code','Section'],values=['Present'],aggfunc=['count','sum'])
print(total)
total.to_excel('dft.xlsx')
import pandas as pd
import xlsxwriter
from datetime import datetime

# Read the CSV file into a DataFrame
df = pd.read_csv('report.csv')
df['Room No']=df['Room No'].str[2:]
current_date = datetime.now().strftime('%d-%m-%Y')
# Create a pivot table with multi-indexing and aggregation
total = pd.pivot_table(df, index=['Room No','Class', 'Subject', 'Section'], values=['Present'], aggfunc=['count','sum'])

# Create a writer object
writer = pd.ExcelWriter('dft.xlsx', engine='xlsxwriter')

# Get the xlsxwriter workbook and worksheet objects
workbook = writer.book
worksheet = workbook.add_worksheet('Sheet1')
# Set the page size (A4 in this example)
worksheet.set_paper(9)  # A4
# Set the left, right, top, and bottom margins (optional)
worksheet.set_margins(left=0.9, right=0.3, top=0.3, bottom=0.3)

# Merge the cells for the header
worksheet.merge_range('A1:G1', 'Half yearly Exam 2023-24 Copy Correction:- '+current_date, workbook.add_format({'bold': True,'font_size': 16, 'align': 'center','border':1, 'valign': 'vcenter'}))
# Define the list of cell references and values
cell_references = ['A2', 'B2', 'C2','D2','E2','F2','G2']
cell_values = ['Class', 'Subject', 'Section','Total','Present','Absent','Remark']

# Write content to the specified cells
for cell_ref, value in zip(cell_references, cell_values):
    worksheet.write(cell_ref, value,workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter','font_size': 14,'border':1}))
# Write the pivot table below the merged header
total.to_excel(writer,sheet_name='Sheet1', startrow=2, header=False, index=True)
last_row_index=len(total)+3
for i in  range(4,last_row_index+1):
    worksheet.write('G' + str(i), '=E'+str(i)+'-F'+str(i))
# Close the4'
# Add 1 to account for zero-based indexing

#print("Last written row index:", last)
writer.close()'''
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
df=pd.read_csv('report.csv')
# Create a DataFrame from the collected data
df = pd.DataFrame(df, columns=["PrNo", "Section", "Class", "Sub_Code", "Subject", "Present", "Room No"])
df['Room No']=df['Room No'].str[2:]
df=df.replace('assRoom',9999)
print(df['Room No'].unique())
df['Room No'] = df['Room No'].astype(int)
print(df['Room No'].unique())
df = df.reset_index(drop=True)
# Create a pivot table with multi-indexing and aggregation
total = pd.pivot_table(df, index=['Room No','Class', 'Subject', 'Section'], values=['Present'], aggfunc=['count','sum'])
#print(sorted_pivot_table)
# Create a new Workbook
workbook = openpyxl.Workbook()

# Access the default sheet (Sheet1)
worksheet = workbook.active

# Set the page size (A4)
worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

# Set the left, right, top, and bottom margins (optional)
worksheet.page_margins.left = 0.9
worksheet.page_margins.right = 0.3
worksheet.page_margins.top = 0.3
worksheet.page_margins.bottom = 0.3

# Merge the cells for the header
worksheet.merge_cells('A1:H1')
header_cell = worksheet['A1']
header_cell.value = 'Half yearly Exam 2023-24 Copy Correction: ' + datetime.now().strftime('%d-%m-%Y')
header_cell.font = Font(bold=True, size=16)
header_cell.alignment = Alignment(horizontal="center", vertical="center")
header_cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

# Define the list of cell references and values for column names
column_names = ['Room','Class', 'Subject', 'Section', 'Total', 'Pre', 'Abs','Remark']

# Write content to the specified cells for column names
for col_idx, column_name in enumerate(column_names, start=1):
    cell = worksheet.cell(row=2, column=col_idx)
    cell.value = column_name
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

# Write the pivot table data below the column names
for row_data in dataframe_to_rows(total, index=True, header=False):
    worksheet.append(row_data)

# Calculate and write the 'Remark' column
for row in range(4, worksheet.max_row + 1):
    worksheet[f'G{row}'] = f'=E{row}-F{row}'
worksheet[f'E{worksheet.max_row +1}'] = f'=SUM(E4:E{worksheet.max_row})'
worksheet[f'F{worksheet.max_row }'] = f'=SUM(F4:F{worksheet.max_row-1})'
worksheet[f'G{worksheet.max_row }'] = f'=E{worksheet.max_row}-F{worksheet.max_row}'
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value == 9999:
            cell.value = 'ClassRoom'  # Set the new value
# Save the workbook
workbook.save('Report22.xlsx')

    



